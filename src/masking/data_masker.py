"""
Módulo para enmascarar datos sensibles en archivos CSV y Excel.
"""

import os
import re
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Optional, Dict, Any

def mask_email(email):
    """
    Enmascara direcciones de email manteniendo el primer carácter y el dominio.
    
    Args:
        email: Email a enmascarar
        
    Returns:
        Email enmascarado
    """
    if pd.isna(email) or "@" not in email:
        return email
    name, domain = email.split("@", 1)
    if not name:
        return email
    return f"{name[0]}{'*' * (len(name)-1)}@{domain}"

def mask_phone(phone):
    """
    Enmascara números de teléfono ocultando los últimos dos dígitos.
    
    Args:
        phone: Número de teléfono a enmascarar
        
    Returns:
        Teléfono enmascarado
    """
    if pd.isna(phone) or not isinstance(phone, str):
        return phone
    return phone[:-2] + "**" if len(phone) > 2 else "**"

def mask_social(url):
    """
    Enmascara URLs de redes sociales ocultando parte del identificador.
    
    Args:
        url: URL de red social a enmascarar
        
    Returns:
        URL enmascarada
    """
    if pd.isna(url) or not isinstance(url, str):
        return url
    return url.split("/")[-1][:2] + "****"

def mask_vowels(text):
    """
    Reemplaza todas las vocales (mayúsculas y minúsculas) en el texto por '*'.
    
    Args:
        text: Texto a enmascarar
        
    Returns:
        Texto con vocales enmascaradas
    """
    if text is None or not isinstance(text, str):
        return text
    return re.sub(r"[aeiouAEIOU]", "*", text)

def mask_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Enmascara datos sensibles en un DataFrame.
    
    Args:
        df: DataFrame a procesar
        
    Returns:
        DataFrame con datos enmascarados
    """
    df_masked = df.copy()
    
    for col in df_masked.columns:
        lower = col.lower()
        if "email" in lower:
            df_masked[col] = df_masked[col].apply(mask_email)
        elif "address" == lower:
            df_masked[col] = df_masked[col].apply(mask_vowels)
        elif "phone" in lower or "tel" in lower:
            df_masked[col] = df_masked[col].apply(mask_phone)
        elif any(s in lower for s in ["facebook", "instagram", "linkedin", "x", "twitter"]):
            df_masked[col] = df_masked[col].apply(mask_social)
            
    return df_masked

def process_csv(file_path: str, output_path: str, modo_prueba: bool = False) -> None:
    """
    Procesa un archivo CSV para enmascarar datos sensibles.
    
    Args:
        file_path: Ruta al archivo CSV de entrada
        output_path: Ruta donde guardar el archivo CSV enmascarado
    """
    df = pd.read_csv(file_path)
    
    # Limitar filas en modo prueba
    if modo_prueba:
        df = df.head(20)
        print(f"🧪 Modo prueba: procesando solo 20 filas")
        
    df = mask_dataframe(df)
    df.to_csv(output_path, index=False)
    print(f"✅ CSV procesado: {os.path.basename(file_path)}")

def process_xlsx(file_path: str, output_path: str, modo_prueba: bool = False) -> None:
    """
    Procesa un archivo Excel para enmascarar datos sensibles.
    
    Args:
        file_path: Ruta al archivo Excel de entrada
        output_path: Ruta donde guardar el archivo Excel enmascarado
    """
    # Verificar si el archivo existe
    if not os.path.exists(file_path):
        print(f"❌ No se encontró el archivo: {file_path}")
        return
        
    # Cargar el workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Procesar solo la hoja llamada "data"
    if 'data' not in wb.sheetnames:
        print(f"❌ No se encontró la hoja 'data' en {file_path}")
        return

    ws = wb['data']
    
    # Leer cabecera para identificar las columnas
    header = [cell.value.lower() if isinstance(cell.value, str) else '' for cell in ws[1]]
    
    # Identificar índices de columnas especiales
    try:
        address_idx = header.index('address')
    except ValueError:
        address_idx = None
        
    email_indices = [i for i, col in enumerate(header) if 'email' in col]
    phone_indices = [i for i, col in enumerate(header) if 'phone' in col or 'tel' in col]
    social_indices = [i for i, col in enumerate(header) 
                     if any(s in col for s in ['facebook', 'instagram', 'linkedin', 'x', 'twitter'])]

    # Limitar filas en modo prueba
    max_row = None
    if modo_prueba:
        max_row = 21  # Fila 1 (cabecera) + 20 filas de datos
        print(f"🧪 Modo prueba: procesando solo 20 filas")
    
    # Procesar cada fila
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Salir si alcanzamos el límite en modo prueba
        if modo_prueba and i > max_row:
            break
        # Enmascarar vocales en la columna 'address'
        if address_idx is not None:
            cell = row[address_idx]
            if isinstance(cell.value, str):
                cell.value = mask_vowels(cell.value)
                
        # Enmascarar emails
        for idx in email_indices:
            cell = row[idx]
            if isinstance(cell.value, str) and '@' in cell.value:
                cell.value = mask_email(cell.value)
                
        # Enmascarar teléfonos
        for idx in phone_indices:
            cell = row[idx]
            if isinstance(cell.value, str) and any(ch.isdigit() for ch in cell.value):
                cell.value = mask_phone(cell.value)
                
        # Enmascarar redes sociales
        for idx in social_indices:
            cell = row[idx]
            if isinstance(cell.value, str) and '/' in cell.value:
                cell.value = mask_social(cell.value)
                
        # Verificar otras celdas que podrían contener datos sensibles
        for cell in row:
            if not isinstance(cell.value, str):
                continue
                
            val = cell.value.lower()
            
            # Detectar emails no identificados por la columna
            if '@' in val and '.' in val and not any(i == cell.column - 1 for i in email_indices):
                cell.value = mask_email(cell.value)
                
            # Detectar teléfonos no identificados por la columna
            elif (any(ch.isdigit() for ch in val) and len(val) >= 7 and 
                  not any(i == cell.column - 1 for i in phone_indices)):
                if sum(1 for ch in val if ch.isdigit()) >= 6:  # Al menos 6 dígitos para ser teléfono
                    cell.value = mask_phone(cell.value)

    # Guardar el archivo procesado
    wb.save(output_path)
    print(f"✅ Excel procesado: {os.path.basename(file_path)}")

def mask_file(file_path: str, output_path: Optional[str] = None, modo_prueba: bool = False) -> None:
    """
    Enmascara datos sensibles en un archivo CSV o Excel.
    
    Args:
        file_path: Ruta al archivo a procesar
        output_path: Ruta opcional donde guardar el resultado
    """
    # Verificar que el archivo existe
    if not os.path.exists(file_path):
        print(f"❌ No se encontró el archivo: {file_path}")
        return
        
    # Determinar tipo de archivo
    file_ext = os.path.splitext(file_path)[1].lower()
    
    # Si no se especifica ruta de salida, crear una por defecto
    if output_path is None:
        dir_name = os.path.dirname(file_path)
        base_name = os.path.basename(file_path)
        name, ext = os.path.splitext(base_name)
        output_path = os.path.join(dir_name, f"{name}_demo{ext}")
    
    # Crear directorio de salida si no existe
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Procesar según tipo de archivo
    if file_ext == '.csv':
        process_csv(file_path, output_path, modo_prueba=modo_prueba)
    elif file_ext in ['.xlsx', '.xls']:
        process_xlsx(file_path, output_path, modo_prueba=modo_prueba)
    else:
        print(f"❌ Formato no soportado: {file_ext}")
        
    return output_path
