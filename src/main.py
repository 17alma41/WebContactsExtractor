"""
Punto de entrada principal para el flujo de trabajo automatizado:
1. Extracción de datos (scraping)
2. Exclusión de emails no deseados
3. Enmascarado de datos sensibles
"""

import os
import sys
import time
import logging
import signal
import pandas as pd
import threading
from pathlib import Path
from typing import List, Dict, Any, Optional
import matplotlib.pyplot as plt
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as XLImage

# Configuración de logging
from src.core.config import (
    BASE_DIR, LOG_DIR, INPUT_DIR, CLEAN_INPUT_DIR, OUTPUT_DIR,
    EXCLUSION_OUTPUT_DIR, DEMO_OUTPUT_DIR, MAX_WORKERS
)

# Importar módulo de visualización
from src.core.visualization import crear_grafico_sectores, crear_grafico_estadisticas

# Configurar logging
logging.basicConfig(
    filename=str(LOG_DIR / "proceso_completo.log"),
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger("main")

# Variable global para control de interrupción
interrupcion_solicitada = threading.Event()


def insertar_imagen_en_excel(ruta_excel: str, ruta_imagen: str) -> None:
    """
    Inserta una imagen en un archivo Excel existente.
    
    Args:
        ruta_excel: Ruta al archivo Excel donde insertar la imagen
        ruta_imagen: Ruta a la imagen a insertar
    """
    try:
        # Cargar el libro de Excel existente
        wb = openpyxl.load_workbook(ruta_excel)
        
        # Seleccionar la hoja activa (primera hoja)
        ws = wb.active
        
        # Crear objeto de imagen
        img = XLImage(ruta_imagen)
        
        # Ajustar tamaño de imagen (opcional)
        img.width = 600  # ancho en píxeles
        img.height = 400  # alto en píxeles
        
        # Insertar imagen en la celda A1
        ws.add_image(img, 'A1')
        
        # Guardar el libro
        wb.save(ruta_excel)
        print(f"✅ Imagen insertada en Excel: {os.path.basename(ruta_excel)}")
    except Exception as e:
        print(f"⚠️ Error al insertar imagen en Excel: {e}")
        logger.error(f"Error al insertar imagen en Excel: {e}")


# Manejador de señales para interrupción controlada
def signal_handler(sig, frame):
    print("\n⏸ Proceso interrumpido. Guardando progreso...")
    interrupcion_solicitada.set()
    # No salir inmediatamente, permitir que el código guarde el progreso
signal.signal(signal.SIGINT, signal_handler)

def limpiar_archivos_csv(modo_prueba: bool = False):
    """
    Ejecuta el proceso de limpieza de archivos CSV.
    
    Args:
        modo_prueba: Si se debe ejecutar en modo prueba (limitado)
    
    Returns:
        bool: True si la limpieza fue exitosa, False en caso contrario
    """
    from src.scraping.csv_cleaner import limpiar_csvs_en_carpeta
    from src.core.data_cleaner import limpiar_archivos_en_carpeta
    
    print("🧹 Limpiando archivos CSV...")
    try:
        # Paso 1: Limpieza básica de CSV
        limpiar_csvs_en_carpeta(
            carpeta_inputs=str(INPUT_DIR),
            carpeta_outputs=str(CLEAN_INPUT_DIR)
        )
        print("✅ Limpieza básica de CSV completada")
        
        # Paso 2: Aplicar transformaciones basadas en configuración
        print("📝 Aplicando transformaciones según configuración...")
        archivos_limpios = limpiar_archivos_en_carpeta(
            carpeta=str(CLEAN_INPUT_DIR),
            carpeta_salida=str(CLEAN_INPUT_DIR)  # Sobrescribir archivos limpios
        )
        
        print(f"✅ Transformaciones completadas: {len(archivos_limpios)} archivos procesados")
        return True
    except Exception as e:
        logger.error(f"Error en limpieza de CSV: {e}")
        print(f"❌ Error en limpieza de CSV: {e}")
        return False

def procesar_scraping(modo_prueba: bool = False):
    """
    Ejecuta el proceso de scraping de datos.
    
    Returns:
        bool: True si el scraping fue exitoso, False en caso contrario
    """
    from src.scraping.scraper import procesar_archivos_csv
    from src.core.excel.column_editor import procesar_csvs_en_carpeta
    
    print("\n🔍 Iniciando proceso de scraping...")
    try:
        archivos_procesados = procesar_archivos_csv(
            carpeta_entrada=str(CLEAN_INPUT_DIR),
            carpeta_salida=str(OUTPUT_DIR),
            max_workers=MAX_WORKERS,
            modo_prueba=modo_prueba
        )
        
        if archivos_procesados:
            print(f"✅ Scraping completado: {len(archivos_procesados)} archivos procesados")
            return True
        else:
            print("⚠️ No se procesaron archivos durante el scraping")
            return False
    except Exception as e:
        logger.error(f"Error en proceso de scraping: {e}")
        print(f"❌ Error en proceso de scraping: {e}")
        return False

def procesar_exclusion_emails(modo_prueba: bool = False):
    """
    Ejecuta el proceso de exclusión de emails no deseados.
    
    Returns:
        bool: True si la exclusión fue exitosa, False en caso contrario
    """
    from src.exclusion.email_exclusion import (
        cargar_exclusiones, procesar_archivo_exclusion, 
        guardar_hojas, guardar_tabla_como_imagen
    )
    from src.core.excel.generator import generar_excel
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    import matplotlib.pyplot as plt
    
    print("\n📧 Iniciando proceso de exclusión de emails...")
    
    try:
        # Cargar exclusiones
        exclusiones = cargar_exclusiones()
        print(f"📋 Cargadas {len(exclusiones)} palabras de exclusión")
        
        # Verificar archivos a procesar
        archivos = [f for f in os.listdir(OUTPUT_DIR) if f.lower().endswith(".xlsx")]
        if not archivos:
            print("⚠️ No se encontraron archivos Excel para procesar")
            return False
            
        print(f"🔄 Procesando {len(archivos)} archivos...")
        
        for fn in archivos:
            entrada = os.path.join(OUTPUT_DIR, fn)
            salida = os.path.join(EXCLUSION_OUTPUT_DIR, fn)
            
            print(f"🔄 Procesando: {fn}")
            
            # Procesar archivo
            if modo_prueba:
                print(f"🧪 Modo prueba activado para exclusión de emails")
            
            # Asegurarse de que el archivo de entrada tenga todas las columnas necesarias
            try:
                hojas_out, estadisticas = procesar_archivo_exclusion(entrada, exclusiones, modo_prueba=modo_prueba)
                
                # Verificar que las hojas necesarias estén presentes
                if "data" not in hojas_out:
                    print("\n⚠️ Advertencia: No se encontró la hoja 'data' en el resultado")
                if "statistics" not in hojas_out:
                    print("\n⚠️ Advertencia: No se encontró la hoja 'statistics' en el resultado")
                    # Crear una hoja de estadísticas básica si no existe
                    hojas_out["statistics"] = pd.DataFrame({
                        "Metric": ["Total rows", "Valid emails", "Excluded emails"],
                        "Value": [len(hojas_out.get("data", pd.DataFrame())), 0, 0]
                    })
                    estadisticas = hojas_out["statistics"]
            except Exception as e:
                print(f"\n❌ Error al procesar archivo para exclusión: {e}")
                # Crear hojas básicas para continuar con el proceso
                df_data = pd.read_excel(entrada, sheet_name="data")
                if modo_prueba:
                    df_data = df_data.head(20)
                hojas_out = {
                    "data": df_data,
                    "statistics": pd.DataFrame({
                        "Metric": ["Total rows", "Valid emails", "Excluded emails"],
                        "Value": [len(df_data), 0, 0]
                    })
                }
                estadisticas = hojas_out["statistics"]
            
            # Guardar hojas
            guardar_hojas(hojas_out, salida)
            
            # Generar imagen de tabla de datos
            if "data" in hojas_out:
                # Ordenar data por reviews si existe la columna
                df_data = hojas_out["data"]
                if "reviews" in df_data.columns:
                    df_data["reviews"] = pd.to_numeric(df_data["reviews"], errors="coerce")
                    df_data = df_data.sort_values("reviews", ascending=False)
                    hojas_out["data"] = df_data
                
                data_path = os.path.join(EXCLUSION_OUTPUT_DIR, fn.replace(".xlsx", "_data.jpg"))
                guardar_tabla_como_imagen(
                    df_data.head(20),
                    data_path,
                    title="Data"
                )
                print(f"\n📷 Imagen de datos guardada: {os.path.basename(data_path)}")
            
            # Generar imagen de sectores si existe la hoja
            df_sectors = hojas_out.get("sectors")
            if df_sectors is not None:
                # ⟪ Columnas que incluyan 'sector'
                sector_cols = [col for col in df_sectors.columns if "sector" in col.lower()]
                # ⟫ Columnas que incluyan 'number' o 'count'
                company_cols = [col for col in df_sectors.columns if any(tok in col.lower() for tok in ("number", "count"))]

                # ⟬ Fallback si solo hay dos columnas
                if not sector_cols and len(df_sectors.columns) == 2:
                    sector_cols = [df_sectors.columns[0]]
                    company_cols = [df_sectors.columns[1]]

                if sector_cols and company_cols:
                    df_sector_imagen = df_sectors[[sector_cols[0], company_cols[0]]].copy()
                    df_sector_imagen.columns = ["Sector", "Number of companies"]
                    df_sector_imagen = df_sector_imagen.sort_values("Number of companies", ascending=False)
                    
                    # Generar imagen de tabla de sectores
                    sectors_path = os.path.join(EXCLUSION_OUTPUT_DIR, fn.replace(".xlsx", "_sectors.jpg"))
                    guardar_tabla_como_imagen(
                        df_sector_imagen.head(20),
                        sectors_path,
                        title="Sectors"
                    )
                    print(f"\n📷 Imagen de tabla de sectores guardada: {os.path.basename(sectors_path)}")
                else:
                    print("\n⚠️ No se encontraron columnas adecuadas en la hoja 'sectors'")
            else:
                print("\n⚠️ Hoja 'sectors' no encontrada")
            
            # Generar imagen gráfica de estadísticas
            try:
                # Usar la función especializada para crear el gráfico de estadísticas
                graph_path = os.path.join(EXCLUSION_OUTPUT_DIR, fn.replace(".xlsx", "_stats.jpg"))
                
                # Crear gráfico con la nueva función
                crear_grafico_estadisticas(
                    estadisticas=estadisticas,
                    titulo=f"Statistics Overview - {os.path.basename(fn)}",
                    ruta_salida=graph_path
                )
                
                print(f"\n📷 Gráfico de estadísticas guardado: {os.path.basename(graph_path)}")
            except Exception as e:
                print(f"\n⚠️ Error al generar gráfico de estadísticas: {e}")
            
            # Insertar imagen en Excel
            wb = load_workbook(salida)
            if "statistics" in wb.sheetnames:
                ws = wb["statistics"]
                img = OpenpyxlImage(graph_path)
                ws.add_image(img, 'A10')
                wb.save(salida)
            
            # Generar tabla data
            df_data = hojas_out.get("data")
            if df_data is not None:
                # Ordenar por reviews si existe la columna
                if "reviews" in df_data.columns:
                    df_data["reviews"] = pd.to_numeric(df_data["reviews"], errors="coerce")
                    df_data = df_data.sort_values("reviews", ascending=False)
                
                # Guardar tabla como imagen
                guardar_tabla_como_imagen(
                    df_data.head(20),
                    os.path.join(EXCLUSION_OUTPUT_DIR, fn.replace(".xlsx", "_data.jpg")),
                    title="Data"
                )
            
            # Generar tabla de sectores
            df_sectors = hojas_out.get("sectors")
            if df_sectors is not None:
                # Columnas que incluyan 'sector'
                sector_cols = [col for col in df_sectors.columns if "sector" in col.lower()]
                # Columnas que incluyan 'number' o 'count'
                company_cols = [col for col in df_sectors.columns 
                               if any(tok in col.lower() for tok in ("number", "count"))]
                
                # Fallback si solo hay dos columnas
                if not sector_cols and len(df_sectors.columns) == 2:
                    sector_cols = [df_sectors.columns[0]]
                    company_cols = [df_sectors.columns[1]]
                
                if sector_cols and company_cols:
                    df_sector_imagen = df_sectors[[sector_cols[0], company_cols[0]]].copy()
                    df_sector_imagen.columns = ["Sector", "Number of companies"]
                    df_sector_imagen = df_sector_imagen.sort_values("Number of companies", ascending=False)
                    guardar_tabla_como_imagen(
                        df_sector_imagen,
                        os.path.join(EXCLUSION_OUTPUT_DIR, fn.replace(".xlsx", "_sectors.jpg")),
                        title="Sectors"
                    )
            
            print(f"✅ Guardado → {salida}")
        
        print("✅ Proceso de exclusión de emails completado")
        return True
    
    except Exception as e:
        logger.error(f"Error en proceso de exclusión de emails: {e}")
        print(f"❌ Error en proceso de exclusión de emails: {e}")
        return False

def procesar_enmascarado(modo_prueba: bool = False):
    """
    Ejecuta el proceso de generación de archivos demo.
    
    Returns:
        bool: True si la generación de archivos demo fue exitosa, False en caso contrario
    """
    from src.masking.data_masker import mask_file
    
    print("\n🎭 Iniciando proceso de generación de archivos demo...")
    
    try:
        # Verificar archivos a procesar
        archivos = [f for f in os.listdir(EXCLUSION_OUTPUT_DIR) 
                   if f.lower().endswith((".xlsx", ".csv"))]
        
        if not archivos:
            print("⚠️ No se encontraron archivos para generar demos")
            return False
            
        print(f"🔄 Generando {len(archivos)} archivos demo...")
        
        for fn in archivos:
            entrada = os.path.join(EXCLUSION_OUTPUT_DIR, fn)
            salida = os.path.join(DEMO_OUTPUT_DIR, fn.replace(".xlsx", "_demo.xlsx").replace(".csv", "_demo.csv"))
            
            print(f"🔄 Generando demo para: {fn}")
            if modo_prueba:
                print(f"🧪 Modo prueba activado para generación de demo")
            
            try:
                # Intentar generar el archivo demo
                mask_file(entrada, salida, modo_prueba=modo_prueba)
                print(f"\n✅ Archivo demo generado → {os.path.basename(salida)}") 
                # Verificar que el archivo de salida tenga todas las hojas necesarias
                if fn.lower().endswith('.xlsx'):
                    try:
                        # Verificar si el archivo tiene todas las hojas necesarias
                        xls = pd.ExcelFile(salida)
                        if 'data' not in xls.sheet_names:
                            print("\n⚠️ Advertencia: No se encontró la hoja 'data' en el archivo demo")
                        if 'statistics' not in xls.sheet_names:
                            print("\n⚠️ Advertencia: No se encontró la hoja 'statistics' en el archivo demo")
                            # Copiar las hojas adicionales del archivo original
                            with pd.ExcelWriter(salida, engine='openpyxl', mode='a') as writer:
                                # Leer hojas del archivo original
                                xls_orig = pd.ExcelFile(entrada)
                                for sheet in xls_orig.sheet_names:
                                    if sheet not in xls.sheet_names and sheet not in ['data']:
                                        df = pd.read_excel(entrada, sheet_name=sheet)
                                        df.to_excel(writer, sheet_name=sheet, index=False)
                    except Exception as e:
                        print(f"\n⚠️ Error al verificar hojas del archivo enmascarado: {e}")
            except Exception as e:
                print(f"\n❌ Error al enmascarar archivo: {e}")
                # Copiar el archivo original como respaldo
                import shutil
                shutil.copy(entrada, salida)
            print(f"✅ Guardado → {salida}")
        
        print("✅ Proceso de enmascarado completado")
        return True
    
    except Exception as e:
        logger.error(f"Error en proceso de enmascarado: {e}")
        print(f"❌ Error en proceso de enmascarado: {e}")
        return False

def ejecutar_flujo_completo(modo_prueba: bool = False):
    """
    Ejecuta el flujo completo de procesamiento de datos.
    
    Args:
        modo_prueba: Si se debe ejecutar en modo prueba (limitado)
    """
    inicio = time.time()
    logger.info("🔄 Inicio del flujo completo de procesamiento.")
    
    # 1. Limpieza de archivos CSV
    if not limpiar_archivos_csv(modo_prueba):
        return
    
    # 2. Proceso de scraping
    if not procesar_scraping(modo_prueba):
        return
    
    # 3. Proceso de exclusión de emails
    if not procesar_exclusion_emails(modo_prueba):
        return
    
    # 4. Proceso de generación de archivos demo
    if not procesar_enmascarado(modo_prueba):
        return
    
    # Mostrar resumen
    duracion = time.time() - inicio
    logger.info(f"✅ Flujo completo terminado en {duracion:.2f}s.")
    print(f"\n✅ Flujo completo terminado en {duracion:.2f}s.")
    
    return True

def ejecutar_procesamiento_automatico(modo_prueba: bool = False):
    """
    Ejecuta el procesamiento automático de archivos sin intervención manual.
    
    Args:
        modo_prueba: Si se debe ejecutar en modo prueba (limitado)
    """
    inicio = time.time()
    logger.info("🔄 Inicio del procesamiento automático.")
    
    # Crear directorios si no existen
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(EXCLUSION_OUTPUT_DIR, exist_ok=True)
    os.makedirs(DEMO_OUTPUT_DIR, exist_ok=True)
    
    # 1. Buscar archivos en data/outputs/
    archivos = [f for f in os.listdir(OUTPUT_DIR) 
               if f.lower().endswith(('.csv', '.xlsx'))]
    
    if not archivos:
        print("\n⚠️ No se encontraron archivos para procesar en", OUTPUT_DIR)
        return False
    
    print(f"\n📂 Encontrados {len(archivos)} archivos para procesar")
    
    # 2. Procesar cada archivo
    for i, archivo in enumerate(archivos, 1):
        print(f"\n[{i}/{len(archivos)}] Procesando: {archivo}")
        
        # Rutas de archivos
        ruta_original = os.path.join(OUTPUT_DIR, archivo)
        ruta_exclusion = os.path.join(EXCLUSION_OUTPUT_DIR, 
                                    archivo.replace('.csv', '.xlsx'))
        ruta_enmascarado = os.path.join(DEMO_OUTPUT_DIR, 
                                      archivo.replace('.csv', '_demo.xlsx')
                                      .replace('.xlsx', '_demo.xlsx'))
        
        # 2.1 Exclusión de emails
        print(f"\n📧 Aplicando exclusión de emails a {archivo}...")
        try:
            from src.exclusion.email_exclusion import (
                cargar_exclusiones, procesar_archivo_exclusion, guardar_hojas
            )
            
            # Cargar exclusiones
            exclusiones = cargar_exclusiones()
            print(f"📋 Cargadas {len(exclusiones)} palabras de exclusión")
            
            # Procesar archivo
            if modo_prueba:
                print(f"🧪 Modo prueba: procesando solo 20 filas")
            
            # Convertir CSV a Excel si es necesario
            if archivo.lower().endswith('.csv'):
                try:
                    df = pd.read_csv(ruta_original)
                    if modo_prueba:
                        df = df.head(20)
                    temp_excel = os.path.join(OUTPUT_DIR, archivo.replace('.csv', '_temp.xlsx'))
                    df.to_excel(temp_excel, sheet_name="data", index=False)
                    ruta_original = temp_excel
                except Exception as e:
                    print(f"\n❌ Error al convertir CSV a Excel: {e}")
                    continue
            
            # Procesar exclusión
            try:
                hojas_out, estadisticas = procesar_archivo_exclusion(
                    ruta_original, exclusiones, modo_prueba=modo_prueba
                )
                
                # Verificar hojas
                if "data" not in hojas_out:
                    print("\n⚠️ No se encontró la hoja 'data' en el resultado")
                    continue
                    
                # Guardar resultado
                guardar_hojas(hojas_out, ruta_exclusion)
                print(f"\n✅ Exclusión completada → {ruta_exclusion}")
                
            except Exception as e:
                print(f"\n❌ Error en exclusión de emails: {e}")
                continue
                
        except Exception as e:
            print(f"\n❌ Error general en exclusión: {e}")
            continue
            
        # 2.2 Generación de archivo demo
        print(f"\n🎭 Generando archivo demo para {os.path.basename(ruta_exclusion)}...")
        try:
            from src.masking.data_masker import mask_file
            
            if modo_prueba:
                print(f"🧪 Modo prueba: procesando solo 20 filas")
                
            # Generar archivo demo
            mask_file(ruta_exclusion, ruta_enmascarado, modo_prueba=modo_prueba)
            print(f"\n✅ Archivo demo generado → {ruta_enmascarado}")
            
        except Exception as e:
            print(f"\n❌ Error al generar archivo demo: {e}")
            continue
    
    # Mostrar resumen
    duracion = time.time() - inicio
    print(f"\n\n✅ Procesamiento automático completado en {duracion:.2f}s")
    print(f"\n📂 Resultados:")
    print(f"  - Archivos originales: {OUTPUT_DIR}")
    print(f"  - Archivos con exclusión: {EXCLUSION_OUTPUT_DIR}")
    print(f"  - Archivos demo: {DEMO_OUTPUT_DIR}")
    
    return True

def main():
    """Función principal del script."""
    print("=" * 60)
    print(" WebContactsExtractor - Flujo Automatizado")
    print("=" * 60)
    print("Opciones disponibles:")
    print("1. Flujo completo (scraping → exclusión → demo)")
    print("2. Solo procesamiento automático (exclusión → demo)")
    print("=" * 60)
    
    opcion = input('Elige una opción (1 o 2): ').strip()
    
    # Selección de modo
    print('\n1 - Modo prueba (20 filas)\n2 - Modo completo')
    modo_prueba = input('Elige (1 o 2): ').strip() == '1'
    
    try:
        if opcion == '1':
            # Flujo completo
            ejecutar_flujo_completo(modo_prueba)
        else:
            # Solo procesamiento automático
            ejecutar_procesamiento_automatico(modo_prueba)
    except KeyboardInterrupt:
        print("\n Proceso cancelado por el usuario.")
        sys.exit(0)

if __name__ == '__main__':
    main()
